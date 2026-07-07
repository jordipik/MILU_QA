import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent

XLSX_DIR = ROOT / "src" / "data" / "excel" / "xlsx"
OUTPUT_DIR = ROOT / "data"
OUTPUT_FILE = OUTPUT_DIR / "EXCEL_FG-FGS.json"

candidates = list(XLSX_DIR.glob("*FG*FGS*.xlsx")) + list(XLSX_DIR.glob("*fg*fgs*.xlsx"))

if not candidates:
    raise FileNotFoundError(f"No he encontrado ningún XLSX FG-FGS en: {XLSX_DIR}")

input_file = candidates[0]

print(f"Leyendo XLSX: {input_file}")

workbook = load_workbook(input_file, data_only=True)
sheet = workbook[workbook.sheetnames[0]]

rows = list(sheet.iter_rows(values_only=True))

if not rows:
    raise ValueError("El Excel está vacío.")

headers = []
for value in rows[0]:
    header = "" if value is None else str(value).strip()
    headers.append(header)

data = []

for row in rows[1:]:
    item = {}

    for index, value in enumerate(row):
        if index >= len(headers):
            continue

        key = headers[index]

        if not key:
            key = f"column_{index + 1}"

        if value is None:
            item[key] = ""
        else:
            item[key] = value

    if any(str(v).strip() for v in item.values()):
        data.append(item)

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

with OUTPUT_FILE.open("w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"OK generado: {OUTPUT_FILE}")
print(f"Filas exportadas: {len(data)}")